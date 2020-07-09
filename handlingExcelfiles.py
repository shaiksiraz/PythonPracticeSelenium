import openpyxl
import xlrd
import xlwt
from openpyxl import workbook
path="C:/Users/shaik/PycharmProjects/seleniumPytest/EmpData.xls"
workbook=xlrd.open_workbook(path)
sheet=workbook.sheet_by_index(0)
rows=sheet.nrows
cols=sheet.ncols
for r in range (0,rows):
    print()
    for c in range(0,cols):
         print(sheet.cell(r,c).value,end="  ")



#writing data in excel files

data =[ [sheet.cell_value(ros, col) for col in range(sheet.ncols)]for ros in range(sheet.nrows)]
print()

print (data)

workbook = xlwt.Workbook()
sheet = workbook.add_sheet('test')

for indexr ,  value in enumerate(data):
    for indexc, value1 in enumerate(value):
        sheet.write(indexr,indexc, value1)

path="C:/Users/shaik/PycharmProjects/seleniumPytest/OutputEmpData.xls"
workbook.save(path)
# for r in range (rows+1 ,rows+5):
#     sheet.cell(r+1,).value=r
#     sheet.cell(r,2).value="siraj-"+r
#     sheet.cell(r,3).value="shaik-"+r
#     sheet.cell(r,4).value=110000+r

# workbook.save(path)